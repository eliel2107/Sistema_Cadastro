import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl

# Configurando a aparência padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()
        
    def layout_config(self):
        self.title("Cadastro")
        self.geometry("1366x768")

    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"])
        self.lb_apm.place(x=1000, y=25)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["light", "Dark", "System"], command=self.change_apm)
        self.opt_apm.place(x=1150, y=75)

    def todo_sistema(self):
        import pathlib
        from openpyxl import Workbook
        from tkcalendar import Calendar, DateEntry
        frame = ctk.CTkFrame(self, width=1400, height=45, corner_radius=0, bg_color="#1E90FF", fg_color="#1E90FF")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Cadastro", font=("century gothic bold", 32), 
                             text_color="#fff").place(x=580, y=8)
        
        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulário!", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"]).place(x=50, y=70)
        

        ficheiro = pathlib.Path("Dados.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = 'Data de Lançamento'
            folha['B1'] = 'Mês'
            folha['C1'] = 'Nota Fiscal'
            folha['D1'] = 'Ordem de Compra'
            folha['E1'] = 'Data de Emissão da NF'
            folha['F1'] = 'Valor da Nota'
            folha['G1'] = 'Volume'
            folha['H1'] = 'Peso'
            folha['I1'] = 'Unidade'
            folha['J1'] = 'Origem'
            folha['K1'] = 'Estado Origem'
            folha['L1'] = 'Destino'
            folha['M1'] = 'Estado Destino'
            folha['N1'] = 'Tipo de Frete'
            folha['O1'] = 'Número da Coleta'
            folha['P1'] = 'Data da Aut. da Coleta'
            folha['Q1'] = 'Transportadora'
            folha['R1'] = 'Situação'
            folha['S1'] = 'Data da Coleta (Transportadora)'
            folha['T1'] = 'Data Prevista da Entrega'
            folha['U1'] = 'Data da Entrega'

            ficheiro.save("Dados.xlsx")  # Salva o arquivo com a extensão correta

        def submit():
            #Pegando os dados dos Entrys
            DataL = DataL_value.get()
            Mes = Mes_combobox.get()
            NotaFiscal = NotaFiscal_value.get()
            Ordem_Compra = Ordem_Compra_value.get()
            Data_EmissaoNF = Data_EmissaoNF_value.get()
            Valor_Nota = Valor_Nota_value.get()
            Volume = Volume_value.get()
            Peso = Peso_value.get()
            Tipo = Tipo_combobox.get()
            Origem = Origem_value.get()
            Estado_Origem = Estado_Origem_combobox.get()
            Destino = Destino_combobox.get()
            Estado_Destino = Estado_Destino_combobox.get()
            TP_Frete = TP_Frete_combobox.get()
            Num_Coleta = Num_Coleta_value.get()
            DT_Autorizada = DT_Autorizada_value.get()
            ID_Transportadora = ID_Transportadora_combobox.get()
            Situacao = Situacao_value.get()
            DT_Coleta = DT_Coleta_value.get()
            DT_Prevista = DT_Prevista_value.get()
            DT_Entrega = DT_Entrega_value.get()

            # Verifica se todos os campos estão preenchidos
            if (DataL == "" or Mes == "" or NotaFiscal == "" or Ordem_Compra == "" or Data_EmissaoNF == "" or
                Valor_Nota == "" or Volume == "" or Peso == "" or Tipo == "" or Origem == "" or Estado_Origem == "" or
                Destino == "" or Estado_Destino == "" or TP_Frete == "" or Num_Coleta == "" or 
                DT_Autorizada == "" or ID_Transportadora == "" or Situacao == ""):
                messagebox.showwarning("Aviso", "Por favor, preencha todos os campos!")
            else:

                ficheiro = openpyxl.load_workbook("Dados.xlsx")
                folha = ficheiro.active

                 # Verifica se a nota fiscal já existe
                for row in folha.iter_rows(min_row=2, max_row=folha.max_row, min_col=3, max_col=3):
                  if row[0].value == NotaFiscal:
                    messagebox.showwarning("Aviso", "Nota Fiscal já existe!")
                    return
                # Adiciona os dados na próxima linha disponível
                nova_linha = folha.max_row + 1
                folha.cell(column=1, row=nova_linha, value=DataL)
                folha.cell(column=2, row=nova_linha, value=Mes)
                folha.cell(column=3, row=nova_linha, value=NotaFiscal)
                folha.cell(column=4, row=nova_linha, value=Ordem_Compra)
                folha.cell(column=5, row=nova_linha, value=Data_EmissaoNF)
                folha.cell(column=6, row=nova_linha, value=Valor_Nota)
                folha.cell(column=7, row=nova_linha, value=Volume)
                folha.cell(column=8, row=nova_linha, value=Peso)
                folha.cell(column=9, row=nova_linha, value=Tipo)
                folha.cell(column=10, row=nova_linha, value=Origem)
                folha.cell(column=11, row=nova_linha, value=Estado_Origem)
                folha.cell(column=12, row=nova_linha, value=Destino)
                folha.cell(column=13, row=nova_linha, value=Estado_Destino)
                folha.cell(column=14, row=nova_linha, value=TP_Frete)
                folha.cell(column=15, row=nova_linha, value=Num_Coleta)
                folha.cell(column=16, row=nova_linha, value=DT_Autorizada)
                folha.cell(column=17, row=nova_linha, value=ID_Transportadora)
                folha.cell(column=18, row=nova_linha, value=Situacao)
                folha.cell(column=19, row=nova_linha, value=DT_Coleta)
                folha.cell(column=20, row=nova_linha, value=DT_Prevista)
                folha.cell(column=21, row=nova_linha, value=DT_Entrega)
                # Salva o arquivo
                ficheiro.save("Dados.xlsx")
                messagebox.showinfo("Sucesso", "Dados inseridos com sucesso!")
                
        def clear():
            DataL_value.set("")
            NotaFiscal_value.set("")
            Ordem_Compra_value.set("")
            Data_EmissaoNF_value.set("")
            Valor_Nota_value.set("")
            Volume_value.set("")
            Peso_value.set("")
            Origem_value.set("")
            Num_Coleta_value.set("")
            DT_Autorizada_value.set("")
            DT_Coleta_value.set("")
            DT_Prevista_value.set("")
            DT_Entrega_value.set("")
            Situacao_value.set("")
            Mes_combobox.set("")
            Tipo_combobox.set("")
            Estado_Origem_combobox.set("")
            Estado_Destino_combobox.set("")
            TP_Frete_combobox.set("")
            ID_Transportadora_combobox.set("")
            Destino_combobox.set("")
            self.update_idletasks()

            # Função para abrir o calendário e preencher o campo de data
        def open_calendar(entry, widget):
            def select_date():
                entry.set(cal.selection_get().strftime('%d-%m-%Y'))
                top.destroy()

            top = Toplevel(self)
            top.transient(self)  # Garante que o popup esteja sempre na frente da janela principal
            top.grab_set()  # Bloqueia a interação com a janela principal enquanto o popup está aberto

            # Obtém a posição do widget de entrada
            x = widget.winfo_rootx() + widget.winfo_width() + 10
            y = widget.winfo_rooty()

            # Define a posição do popup do calendário
            top.geometry(f"+{x}+{y}")

            cal = Calendar(top, selectmode='day', date_pattern='dd-mm-yyyy')
            cal.pack(pady=20)
            Button(top, text="Selecionar", command=select_date).pack()
        
        #Textos de Variaveis
        DataL_value = StringVar()
        NotaFiscal_value = StringVar()
        Ordem_Compra_value = StringVar()
        Data_EmissaoNF_value = StringVar()
        Valor_Nota_value = StringVar()
        Volume_value = StringVar()
        Peso_value = StringVar()
        Origem_value = StringVar()
        Num_Coleta_value = StringVar()
        DT_Autorizada_value = StringVar()
        Situacao_value = StringVar()
        DT_Coleta_value = StringVar()
        DT_Prevista_value = StringVar()
        DT_Entrega_value = StringVar()
        Mes_combobox = StringVar()

        #Entrys
        DataL_entry = ctk.CTkEntry(self, width=100, textvariable=DataL_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        NotaFiscal_entry = ctk.CTkEntry(self, width=100, textvariable=NotaFiscal_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Ordem_Compra_entry = ctk.CTkEntry(self, width=150, textvariable=Ordem_Compra_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Data_EmissaoNF_entry = ctk.CTkEntry(self, width=150, textvariable=Data_EmissaoNF_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Valor_Nota_entry = ctk.CTkEntry(self, width=150, textvariable=Valor_Nota_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Volume_entry = ctk.CTkEntry(self, width=100, textvariable=Volume_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Peso_entry = ctk.CTkEntry(self, width=100, textvariable=Peso_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Origem_entry = ctk.CTkEntry(self, width=450, textvariable=Origem_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Num_Coleta_entry = ctk.CTkEntry(self, width=150, textvariable=Num_Coleta_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        DT_Autorizada_entry = ctk.CTkEntry(self, width=100, textvariable=DT_Autorizada_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        Situacao_entry = ctk.CTkEntry(self, width=150, textvariable=Situacao_value, font=("Century Gothic", 16), fg_color="transparent", justify="center")
        DT_Coleta_entry = ctk.CTkEntry(self, width=100, textvariable=DT_Coleta_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        DT_Prevista_entry = ctk.CTkEntry(self, width=100, textvariable=DT_Prevista_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")
        DT_Entrega_entry = ctk.CTkEntry(self, width=100, textvariable=DT_Entrega_value, font=("Century Gothic ", 16), fg_color="transparent", justify="center")

        #ComboBox
        Tipo_combobox = ctk.CTkComboBox(self, values=["G", "KG"], font=("Century Gothic bold", 14), state="readonly", width=75, justify="center")
        

        Mes_combobox = ctk.CTkComboBox(self, values=["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"], font=("Century Gothic bold", 14), state="readonly", justify="center")
        
        Estado_Origem_combobox = ctk.CTkComboBox(self, values=["AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"], font=("Century Gothic bold", 14), 
        state="readonly", justify="center", width=75)
        

        Destino_combobox = ctk.CTkComboBox(self,values=[
            "ESOM BASE MOSSORÓ",
            "CHN IMPORTAÇÃO DE MÁQUINAS E EQUIPAMENTOS LTDA",
            "CONAUT CONTROLES AUTOMÁTICOS LTDA",
            "CONEX ELETROMECÂNICA INDÚSTRIA E COMÉRCIO LTDA",
            "DHARMACOM TELECOMUNICAÇÕES LTDA",
            "EAGLEBURGMANN DO BRASIL VEDAÇÕES INDUSTRIAIS LTDA",
            "EMERSON PROCESS MANAGEMENT LTDA",
            "ESOM BASE ALCOBAÇA (PRADO)",
            "ESOM BASE ARACRUZ",
            "ESOM BASE ATALAIA",
            "ESOM BASE CAMACARI",
            "ESOM BASE CAMAÇARI",
            "ESOM BASE CATU",
            "ESOM BASE CENTRO",
            "ESOM BASE COARI",
            "ESOM BASE ITABUNA",
            "ESOM BASE JABOATÃO DOS GUARARAPES",
            "ESOM BASE JOÃO PESSOA",
            "ESOM BASE MACAÍBA",
            "ESOM BASE MANAUS",
            "ESOM BASE MARACANAÚ",
            "ESOM BASE MOSSORÓ",
            "ESOM BASE OLARIA",
            "ESOM BASE PILAR",
            "ESOM BASE PIUMA",
            "ESOM BASE VITÓRIA",
            "ICONE TECNOLOGIA E AUTOMAÇÃO LTDA",
            "INSTITUTO DE PESQUISAS TECNOLÓGICAS DO ESTADO DE SÃO PAULO S.A",
            "IOPE INSTRUMENTOS DE PRECISÃO LTDA",
            "ITUFLUX INSTRUMENTOS DE MEDIÇÃO LTDA",
            "KOCH METALURGICA S.A",
            "MEC Q COMÉRCIO E SERVIÇOS DE METROLOGIA INDUSTRIAL LTDA",
            "NT TRANSFORMEDORES COMÉRCIO DE MATERIAL ELÉTRICO LTDA",
            "POWERENG SISTEMAS DE ENERGIA LTDA",
            "PRECISÃO SERVIÇOS TÉCNICOS ESPECIALIZADOS LTDA",
            "PRESYS INSTRUMENTOS E SISTEMAS LTDA",
            "SCHWEITZER ENGINEERING LABORATORIES COMERCIAL LTDA",
            "SERVENTEC COMÉRCIO - INDÚSTRIA E SERVIÇOS DE MANUTENÇÃO INDUSTRIAL",
            "SERVIÇO NACIONAL DE APRENDIZAGEM INDUSTRIAL - CENTRO DE TECNOLOGIAS DO GÁS E ENERGIAS RENOVÁVEIS (CTGÁS)",
            "SONDEQ INDÚSTRIA DE SONDAS E EQUIPAMENTOS LTDA",
            "TAG - BASE ALCOBACA (PRADO)",
            "TAG - BASE ARACRUZ (ESOM ARACRUZ)",
            "TAG - BASE ARACRUZ (ESOM BASE ARACRUZ)",
            "TAG - BASE ATALAIA (ESOM ATALAIA)",
            "TAG - BASE ATALAIA (ESOM BASE ATALAIA)",
            "TAG - BASE CATU (ESOM CATU)",
            "TAG - BASE JABOATÃO DOS GUARARAPES (ESOM JABOATÃO DOS GUARARAPES)",
            "TAG - BASE MANAUS (ESOM BASE MANAUS)",
            "TAG - BASE PIUMA (ESOM PIUMA)",
            "TECNOISO TECNOLOGIA E SOLUÇÕES INDUSTRIAIS LTDA",
            "TEREME TÉCNICA DE RECUPERAÇÃO DE MÁQUINA ELÉTRICA LTDA",
            "UNIS GROUP REPARAÇÃO ELETRÔNICA INDUSTRIAL",
            "VANASA MULTIGÁS ENGENHARIA INDÚSTRIA E COMÉRCIO LTDA"
            ], 
            font=("Century Gothic bold", 14),
            width=450, justify="center")
        
        Estado_Destino_combobox = ctk.CTkComboBox(self, values=["AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"], font=("Century Gothic bold", 14), state="readonly",
        width=75, justify="center")

        TP_Frete_combobox = ctk.CTkComboBox(self, values=["RODOVIÁRIO CONVENCIONAL","AÉREO CONVENCIONAL","RODOVIÁRIO EMERGENCIAL", "AÉREO EMERGENCIAL"], 
        font=("Century Gothic bold", 14), state="readonly", width=240, justify="center")

        ID_Transportadora_combobox = ctk.CTkComboBox(self, values=["LAST MILE", "ATUAL CARGAS"], font=("Century Gothic bold", 14), state="readonly", width=240, justify="center")

        #Labels
        lb_DataL = ctk.CTkLabel(self, text="Data de Lançamento", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Mes = ctk.CTkLabel(self, text="Mês", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_NotaFiscal = ctk.CTkLabel(self, text="Nota Fiscal", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Ordem_Compra = ctk.CTkLabel(self, text="Ordem de Compra", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Data_EmissaoNF = ctk.CTkLabel(self, text="Data de Emissão NF", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Valor_Nota = ctk.CTkLabel(self, text="Valor da Nota", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Volume = ctk.CTkLabel(self, text="Volume", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Peso = ctk.CTkLabel(self, text="Peso", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Tipo = ctk.CTkLabel(self, text="Unidade", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Origem = ctk.CTkLabel(self, text="Origem", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Estado_Origem = ctk.CTkLabel(self, text="UF Origem", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Destino = ctk.CTkLabel(self, text="Destino", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Estado_Destino = ctk.CTkLabel(self, text="UF Destino", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_TP_Frete = ctk.CTkLabel(self, text="Tipo de Frete", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Num_Coleta = ctk.CTkLabel(self, text="Número de Coleta", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_DT_Autorizada = ctk.CTkLabel(self, text="Data Autorizada", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_ID_Transportadora = ctk.CTkLabel(self, text="Transportadora", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_Situacao = ctk.CTkLabel(self, text="Situação", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_DT_Coleta = ctk.CTkLabel(self, text="Data da Coleta", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_DT_Prevista = ctk.CTkLabel(self, text="Data Prevista da Entrega", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        lb_DT_Entrega = ctk.CTkLabel(self, text="Data da Entrega", font=("century gothic bold", 16)\
                             , text_color=["#000", "#fff"], anchor="center")
        # Botões para abrir o calendário
        DataL_calendar_btn = ctk.CTkButton(self, text="📅", width=30, command=lambda: open_calendar(DataL_value, DataL_entry))
        Data_EmissaoNF_calendar_btn = ctk.CTkButton(self, text="📅", width=30, command=lambda: open_calendar(Data_EmissaoNF_value, Data_EmissaoNF_entry))
        DT_Autorizada_calendar_btn = ctk.CTkButton(self, text="📅", width=30, command=lambda: open_calendar(DT_Autorizada_value, DT_Autorizada_entry))
        DT_Coleta_calendar_btn = ctk.CTkButton(self, text="📅", width=30, command=lambda: open_calendar(DT_Coleta_value, DT_Coleta_entry))
        DT_Prevista_calendar_btn = ctk.CTkButton(self, text="📅", width=30, command=lambda: open_calendar(DT_Prevista_value, DT_Prevista_entry))
        DT_Entrega_calendar_btn = ctk.CTkButton(self, text="📅", width=30, command=lambda: open_calendar(DT_Entrega_value, DT_Entrega_entry))

        #Posicionando os elementos na tela
        lb_DataL.place(x=50, y=120)
        DataL_entry.place(x=50, y=150)
        DataL_calendar_btn.place(x=155, y=150)

        lb_Mes.place(x=220, y=120) 
        Mes_combobox.place(x=220, y=150) 

        lb_NotaFiscal.place(x=400, y=120)
        NotaFiscal_entry.place(x=400, y=150)

        lb_Ordem_Compra.place(x=520, y=120)
        Ordem_Compra_entry.place(x=520, y=150)

        lb_Data_EmissaoNF.place (x=50, y=190)
        Data_EmissaoNF_entry.place(x=50, y=220)
        Data_EmissaoNF_calendar_btn.place(x=205, y=220)

        lb_Valor_Nota.place (x=690, y=120)
        Valor_Nota_entry.place(x=690, y=150)

        lb_Volume.place(x=260, y=190)
        Volume_entry.place(x=260, y=220)

        lb_Peso.place(x=400, y=190)
        Peso_entry.place(x=400, y=220)

        lb_Tipo.place(x=520, y=190)
        Tipo_combobox.place(x=520, y=220)

        lb_Origem.place(x=50, y=265)
        Origem_entry.place(x=50, y=295)

        lb_Estado_Origem.place(x=520, y=265)
        Estado_Origem_combobox.place(x=520, y=295)

        lb_Destino.place(x=50, y=405)
        Destino_combobox.place(x=50, y=435)

        lb_Estado_Destino.place(x=520, y=405)
        Estado_Destino_combobox.place(x=520, y=435)

        lb_TP_Frete.place(x=50, y=335)
        TP_Frete_combobox.place(x=50, y=365)

        lb_Num_Coleta.place(x=580, y=335)
        Num_Coleta_entry.place(x=580, y=365) 

        lb_DT_Autorizada.place(x=755, y=335)
        DT_Autorizada_entry.place(x=755, y=365)
        DT_Autorizada_calendar_btn.place(x=860, y=365)

        lb_ID_Transportadora.place(x=310, y=335) #Irá ser mudado depois
        ID_Transportadora_combobox.place(x=310, y=365)

        lb_Situacao.place(x=615, y=405)
        Situacao_entry.place(x=615, y=435)

        lb_DT_Coleta.place(x=785, y=405)
        DT_Coleta_entry.place(x=785, y=435)
        DT_Coleta_calendar_btn.place(x=890, y=435)

        lb_DT_Prevista.place(x=905, y=335)
        DT_Prevista_entry.place(x=905, y=365)
        DT_Prevista_calendar_btn.place(x=1010, y=365)

        lb_DT_Entrega.place(x=935, y=405)
        DT_Entrega_entry.place(x=935, y=435)
        DT_Entrega_calendar_btn.place(x=1040, y=435)

        btn_submit = ctk.CTkButton(self, text="Salvar Dados".upper(), command=submit, fg_color="#151", 
                                   hover_color="#131").place(x=400, y=550)

        btn_submit = ctk.CTkButton(self, text="Limpar Campos".upper(), command=clear, fg_color="#555", 
                                   hover_color="#333").place(x=600, y=550)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__ == "__main__":
    app = App()
    app.mainloop()

